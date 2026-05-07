"""Create XLSX tool."""
from typing import Any, Dict
from .base import Tool, ToolResult


class CreateXlsxTool(Tool):
    """Tool for creating Excel spreadsheets."""

    def __init__(self):
        super().__init__(
            name="create_xlsx",
            description="Create an Excel file. Input: JSON with 'data' (array of rows), 'headers' (optional), 'output' (filename).",
        )

    async def execute(self, **kwargs) -> ToolResult:
        try:
            from openpyxl import Workbook
            
            data = kwargs.get("data", [])
            headers = kwargs.get("headers", [])
            output = kwargs.get("output", "spreadsheet.xlsx")
            
            wb = Workbook()
            ws = wb.active
            
            if headers:
                for col, header in enumerate(headers, 1):
                    ws.cell(1, col, header)
                start_row = 2
            else:
                start_row = 1
            
            for row_idx, row in enumerate(data, start_row):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row_idx, col_idx, value)
            
            wb.save(output)
            return ToolResult(success=True, result=f"Created {output} with {len(data)} rows")
        except ImportError:
            return ToolResult(success=False, result=None, error="openpyxl not installed. Run: pip install openpyxl")
        except Exception as e:
            return ToolResult(success=False, result=None, error=str(e))

    def _get_parameters_schema(self) -> Dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "data": {"type": "array", "description": "Array of rows (each row is an array of values)"},
                "headers": {"type": "array", "description": "Optional column headers"},
                "output": {"type": "string", "description": "Output filename (default: spreadsheet.xlsx)"}
            },
            "required": ["data"]
        }