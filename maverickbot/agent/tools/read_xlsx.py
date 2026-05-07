"""Read XLSX tool."""
from typing import Any, Dict
from .base import Tool, ToolResult


class ReadXlsxTool(Tool):
    """Tool for reading Excel spreadsheets."""

    def __init__(self):
        super().__init__(
            name="read_xlsx",
            description="Read an Excel file. Input: JSON with 'file' (path), 'sheet' (optional sheet name).",
        )

    async def execute(self, **kwargs) -> ToolResult:
        try:
            from openpyxl import load_workbook
            
            file_path = kwargs.get("file", "")
            sheet_name = kwargs.get("sheet", None)
            
            if not file_path:
                return ToolResult(success=False, result=None, error="No file specified")
            
            wb = load_workbook(file_path, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active
            
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else "" for cell in row])
            
            result = f"Sheet: {ws.title}\nRows: {len(rows)}\n\n"
            result += "\n".join([str(row) for row in rows[:50]])
            
            return ToolResult(success=True, result=result)
        except ImportError:
            return ToolResult(success=False, result=None, error="openpyxl not installed. Run: pip install openpyxl")
        except FileNotFoundError:
            return ToolResult(success=False, result=None, error=f"File not found: {file_path}")
        except Exception as e:
            return ToolResult(success=False, result=None, error=str(e))

    def _get_parameters_schema(self) -> Dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "file": {"type": "string", "description": "Path to the XLSX file"},
                "sheet": {"type": "string", "description": "Sheet name (optional)"}
            },
            "required": ["file"]
        }